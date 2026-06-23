"""CLI for GatePassX — Dinner & Event Gate Pass tooling."""

import json
import csv
import os
import sys
from datetime import datetime, date
from pathlib import Path
from typing import List, Optional

import click

from .models import (
    GatePass, PassCategory, EventType, QRAlignment, OverlayText,
    PassDesignSettings, get_qr_secret,
)
from .generator import (
    create_pass_pdf,
    generate_batch_pdfs,
    create_passes_sheet,
    save_qr_image,
    export_batch_images,
    export_template_as_image,
    lock_input_file,
    check_input_lock,
    unlock_input_file,
)
from .models import PassManifest
from .utils import verify_qr_payload

__version__ = "1.0.0"


def _load_passes_from_file(path: Path) -> List[GatePass]:
    """Load passes from JSON, YAML, CSV, or XLSX."""
    suffix = path.suffix.lower()

    if suffix == ".json":
        data = json.loads(path.read_text())
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            items = data.get("passes", [data]) if "pass_id" not in data else [data]
        else:
            items = []
        return [GatePass(**_coerce_item(item)) for item in items]

    elif suffix in {".yaml", ".yml"}:
        import yaml
        data = yaml.safe_load(path.read_text())
        items = data if isinstance(data, list) else data.get("passes", [])
        return [GatePass(**_coerce_item(item)) for item in items]

    elif suffix == ".csv":
        with path.open(newline="") as f:
            reader = csv.DictReader(f)
            passes = []
            for row in reader:
                passes.append(GatePass(**_coerce_csv_row(row)))
            return passes

    elif suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
        passes = []
        for cells in rows_iter:
            row = {headers[i]: cells[i] for i in range(min(len(headers), len(cells))) if headers[i]}
            if not any(row.values()):
                continue
            passes.append(GatePass(**_coerce_xlsx_row(row)))
        wb.close()
        return passes

    else:
        raise click.ClickException(f"Unsupported input format: {suffix}")


def _coerce_item(item: dict) -> dict:
    """Normalize field types for a JSON/YAML pass dict."""
    if "category" in item and isinstance(item["category"], str):
        item["category"] = PassCategory(item["category"].upper())
    if "event_type" in item and isinstance(item["event_type"], str):
        item["event_type"] = EventType(item["event_type"].upper())
    # Legacy field mapping: operator → organizer, trip_type → event_type
    if "operator" in item and "organizer" not in item:
        item["organizer"] = item.pop("operator")
    if "trip_type" in item and "event_type" not in item:
        mapping = {"HAJJ": "OTHER", "UMRAH": "OTHER"}
        item["event_type"] = EventType(mapping.get(item["trip_type"].upper(), item["trip_type"].upper()))
        del item["trip_type"]
    return item


def _coerce_csv_row(row: dict) -> dict:
    """Normalize a CSV row into GatePass-compatible types."""
    if "valid_from" in row and row["valid_from"]:
        row["valid_from"] = date.fromisoformat(row["valid_from"])
    if "valid_to" in row and row["valid_to"]:
        row["valid_to"] = date.fromisoformat(row["valid_to"])
    if "issued_at" in row and row["issued_at"]:
        row["issued_at"] = datetime.fromisoformat(row["issued_at"])
    if "category" in row and row["category"]:
        row["category"] = PassCategory(row["category"].upper())
    if "event_type" in row and row["event_type"]:
        row["event_type"] = EventType(row["event_type"].upper())
    # Legacy columns
    if "operator" in row and "organizer" not in row:
        row["organizer"] = row.pop("operator", "")
    if "trip_type" in row and "event_type" not in row:
        tt = row.pop("trip_type", "")
        if tt:
            row["event_type"] = EventType.OTHER
    return row


def _coerce_xlsx_row(row: dict) -> dict:
    """Normalize an XLSX row into GatePass-compatible types.

    openpyxl returns datetime.datetime for date cells, so handle those
    before stringifying everything else.
    """
    if "valid_from" in row and row["valid_from"]:
        v = row["valid_from"]
        if isinstance(v, datetime):
            row["valid_from"] = v.date()
        elif isinstance(v, date):
            pass
        else:
            row["valid_from"] = date.fromisoformat(str(v)[:10])
    if "valid_to" in row and row["valid_to"]:
        v = row["valid_to"]
        if isinstance(v, datetime):
            row["valid_to"] = v.date()
        elif isinstance(v, date):
            pass
        else:
            row["valid_to"] = date.fromisoformat(str(v)[:10])
    if "issued_at" in row and row["issued_at"]:
        v = row["issued_at"]
        if isinstance(v, datetime):
            pass
        else:
            row["issued_at"] = datetime.fromisoformat(str(v))
    if "category" in row and row["category"]:
        row["category"] = PassCategory(str(row["category"]).upper())
    if "event_type" in row and row["event_type"]:
        row["event_type"] = EventType(str(row["event_type"]).upper())
    for key in list(row.keys()):
        v = row[key]
        if v is None:
            row[key] = ""
        elif isinstance(v, (date, datetime, PassCategory, EventType)):
            pass
        elif not isinstance(v, str):
            row[key] = str(v)
    if "operator" in row and "organizer" not in row:
        row["organizer"] = row.pop("operator", "")
    if "trip_type" in row and "event_type" not in row:
        tt = row.pop("trip_type", "")
        if tt:
            row["event_type"] = EventType.OTHER
    return row


def _build_design_settings(
    qr_size: float,
    qr_align: str,
    qr_border: int,
    qr_transparent: bool,
    overlay_texts: tuple,
    text_sizes: tuple,
) -> PassDesignSettings:
    """Build PassDesignSettings from CLI arguments."""
    overlays = []
    for i, txt in enumerate(overlay_texts):
        if txt:
            sz = text_sizes[i] if i < len(text_sizes) else 10.0
            overlays.append(OverlayText(text=txt, size=sz))
    return PassDesignSettings(
        qr_size_mm=qr_size,
        qr_alignment=QRAlignment(qr_align),
        qr_border=qr_border,
        qr_transparent_bg=qr_transparent,
        overlay_texts=overlays,
    )


# ── CLI Group ────────────────────────────────────────────────────────────────

@click.group()
@click.version_option(__version__, prog_name="gatepassx")
def cli():
    """GatePassX — Dinner & Event Gate Pass Generator and Scanner CLI.

    Generate printable PDFs with QR codes, create/validate passes,
    and manage event gate pass data from the command line.
    """
    pass


# ── generate ─────────────────────────────────────────────────────────────────

@cli.command()
@click.option("-i", "--input", "input_path", required=True,
              type=click.Path(exists=True, path_type=Path),
              help="JSON / YAML / CSV / XLSX file containing pass data.")
@click.option("-o", "--out", "output_dir", default="generated_passes",
              type=click.Path(path_type=Path),
              help="Output directory for PDFs (default: generated_passes/).")
@click.option("--secret", default=None, help="QR signing secret (default: $GATEPASSX_QR_SECRET).")
@click.option("--sheet", is_flag=True, help="Also produce a batch summary sheet PDF.")
@click.option("--qr-only", is_flag=True, help="Generate QR PNG images instead of PDFs.")
@click.option("--qr-size", default=40.0, type=float, help="QR code size in mm (default: 40).")
@click.option("--qr-align", default="right",
              type=click.Choice(["left", "center", "right"], case_sensitive=False),
              help="QR code horizontal alignment (default: right).")
@click.option("--qr-border", default=0, type=int,
              help="QR quiet-zone border modules, 0 = no white border (default: 0).")
@click.option("--qr-transparent/--qr-white-bg", default=True,
              help="Use transparent QR background (default: transparent).")
@click.option("--overlay-text", "overlay_texts", multiple=True, default=(),
              help="Optional text line to display on the pass (repeatable).")
@click.option("--text-size", "text_sizes", multiple=True, default=(), type=float,
              help="Font size for corresponding overlay text (repeatable, default: 10).")
@click.option("--template", "template_path", default=None,
              type=click.Path(exists=True, path_type=Path),
              help="Template PDF to use as background (overlay pass data on template).")
@click.option("--png", "export_png", is_flag=True,
              help="Export each pass as a high-resolution PNG image.")
@click.option("--jpeg", "export_jpeg", is_flag=True,
              help="Export each pass as a high-resolution JPEG image.")
@click.option("--dpi", default=300, type=int,
              help="DPI for PNG/JPEG export (default: 300).")
@click.option("--jpeg-quality", default=95, type=int,
              help="JPEG quality 1-100 (default: 95).")
@click.option("--force", is_flag=True,
              help="Override lock protection and regenerate even if data was altered.")
@click.option("--no-lock", is_flag=True,
              help="Skip creating a lock file after generation.")
def generate(input_path: Path, output_dir: Path, secret: Optional[str], sheet: bool, qr_only: bool,
             qr_size: float, qr_align: str, qr_border: int, qr_transparent: bool,
             overlay_texts: tuple, text_sizes: tuple, template_path: Optional[Path],
             export_png: bool, export_jpeg: bool, dpi: int, jpeg_quality: int,
             force: bool, no_lock: bool):
    """Generate gate pass PDFs, QR images, or high-res PNG/JPEG from a data file.

    After generation the input file is locked (.lock manifest). Any subsequent
    attempt to generate from an altered file is blocked unless --force is used.
    """
    # ── Lock check ──────────────────────────────────────────────────────
    manifest, intact = check_input_lock(str(input_path))
    if manifest is not None and not intact:
        click.echo(f"✗ LOCKED: {input_path} has been altered since last generation.")
        click.echo(f"  Last generated: {manifest.generated_at.strftime('%Y-%m-%d %H:%M UTC')}")
        click.echo(f"  Passes: {manifest.pass_count}  →  {manifest.output_dir}")
        if not force:
            click.echo(f"  Use --force to override, or 'gatepassx unlock {input_path}' to remove lock.")
            sys.exit(1)
        click.echo("  --force: overriding lock protection.")

    if secret is None:
        secret = get_qr_secret()

    design = _build_design_settings(qr_size, qr_align, qr_border, qr_transparent, overlay_texts, text_sizes)
    if template_path:
        design.template_path = str(template_path)

    passes = _load_passes_from_file(input_path)
    click.echo(f"✓ Loaded {len(passes)} pass(es) from {input_path}")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if qr_only:
        qr_dir = output_dir / "qr_codes"
        qr_dir.mkdir(exist_ok=True)
        for p in passes:
            p.compute_qr_payload(secret=secret)
            safe_id = p.pass_id.replace("/", "-").replace(" ", "_")
            save_qr_image(p.qr_payload, str(qr_dir / f"{safe_id}.png"),
                          transparent_bg=design.qr_transparent_bg, border=design.qr_border)
        click.echo(f"✓ Generated {len(passes)} QR code(s) in {qr_dir}")
    else:
        created = generate_batch_pdfs(passes, str(output_dir), secret=secret, design=design)
        click.echo(f"✓ Generated {len(created)} PDF(s) in {output_dir}")

    if export_png:
        png_dir = output_dir / "png"
        png_paths = export_batch_images(passes, str(png_dir), secret=secret, design=design,
                                        dpi=dpi, fmt="png")
        click.echo(f"✓ Exported {len(png_paths)} PNG(s) at {dpi} DPI → {png_dir}")

    if export_jpeg:
        jpg_dir = output_dir / "jpeg"
        jpg_paths = export_batch_images(passes, str(jpg_dir), secret=secret, design=design,
                                        dpi=dpi, fmt="jpeg", jpeg_quality=jpeg_quality)
        click.echo(f"✓ Exported {len(jpg_paths)} JPEG(s) at {dpi} DPI → {jpg_dir}")

    if sheet:
        sheet_path = output_dir / "batch_sheet.pdf"
        create_passes_sheet(passes, str(sheet_path), secret=secret, design=design)
        click.echo(f"✓ Created batch summary sheet: {sheet_path}")

    # ── Auto-lock ───────────────────────────────────────────────────────
    if not no_lock:
        m = lock_input_file(str(input_path), str(output_dir), len(passes))
        click.echo(f"🔒 Locked {input_path.name} ({len(passes)} passes, hash {m.input_hash[:12]}…)")

    click.echo("Done.")


# ── new ──────────────────────────────────────────────────────────────────────

@cli.command()
@click.option("-o", "--out", default=None, type=click.Path(path_type=Path),
              help="Output JSON file (default: stdout).")
@click.option("--event-name", default="Annual Dinner 2026", help="Event name.")
@click.option("--event-type", default="DINNER",
              type=click.Choice([e.value for e in EventType], case_sensitive=False),
              help="Event type.")
def new(out: Optional[Path], event_name: str, event_type: str):
    """Interactively create a new gate pass (outputs JSON)."""
    click.echo("── GatePassX: New Pass ──")

    et = EventType(event_type.upper())
    ts = datetime.utcnow().strftime("%H%M%S")
    default_id = f"GPX-{et.value}-{datetime.utcnow().year}-{ts}"

    data = {}
    data["pass_id"] = click.prompt("Pass ID", default=default_id)
    data["event_name"] = click.prompt("Event name", default=event_name)
    data["event_type"] = et
    data["full_name"] = click.prompt("Guest full name")
    data["id_number"] = click.prompt("ID / Registration number")
    data["phone"] = click.prompt("Phone", default="") or None
    data["email"] = click.prompt("Email", default="") or None
    data["organizer"] = click.prompt("Organizer", default="Event Organizer")
    data["category"] = click.prompt(
        "Category",
        type=click.Choice([c.value for c in PassCategory], case_sensitive=False),
        default="GUEST",
    )
    data["valid_from"] = click.prompt("Valid from (YYYY-MM-DD)", default=date.today().isoformat())
    data["valid_to"] = click.prompt("Valid to (YYYY-MM-DD)",
                                    default=(date.today().replace(day=min(date.today().day + 3, 28))).isoformat())
    data["gate"] = click.prompt("Gate / Entrance", default="Main Gate")
    data["table_number"] = click.prompt("Table / Seat number", default="") or None
    data["group_ref"] = click.prompt("Group / Invite reference", default="") or None
    data["issued_by"] = click.prompt("Issued by", default="GatePassX CLI")

    gp = GatePass(**data)
    gp.compute_qr_payload(secret=get_qr_secret())
    result = gp.model_dump_json(indent=2)

    if out:
        out.write_text(result)
        click.echo(f"✓ Wrote pass to {out}")
        click.echo(f"  Run: gatepassx generate -i {out} -o ./passes --sheet")
    else:
        click.echo("\n── Pass JSON ──")
        click.echo(result)


# ── validate ─────────────────────────────────────────────────────────────────

@cli.command()
@click.argument("pass_file", type=click.Path(exists=True, path_type=Path))
@click.option("--secret", default=None, help="QR signing secret for verification.")
def validate(pass_file: Path, secret: Optional[str]):
    """Validate and display details of a pass JSON file."""
    if secret is None:
        secret = get_qr_secret()

    data = json.loads(pass_file.read_text())
    gp = GatePass(**data)
    payload = gp.compute_qr_payload(secret=secret)

    click.echo(f"  Pass ID:    {gp.pass_id}")
    click.echo(f"  Event:      {gp.event_name} ({gp.event_type})")
    click.echo(f"  Guest:      {gp.full_name}")
    click.echo(f"  Category:   {gp.category}")
    click.echo(f"  Organizer:  {gp.organizer}")
    click.echo(f"  Valid:      {gp.valid_from} → {gp.valid_to}")
    click.echo(f"  Gate:       {gp.gate or '—'}")
    click.echo(f"  Table:      {gp.table_number or '—'}")
    click.echo(f"\n  QR payload ({len(payload)} bytes):")
    click.echo(f"  {payload}")

    if gp.qr_payload:
        valid, qr_data = verify_qr_payload(gp.qr_payload, secret=secret)
        status = "✓ VALID" if valid else "✗ INVALID"
        click.echo(f"\n  Signature:  {status}")


# ── scan ─────────────────────────────────────────────────────────────────────

@cli.command()
@click.argument("qr_payload")
@click.option("--secret", default=None, help="QR signing secret for verification.")
def scan(qr_payload: str, secret: Optional[str]):
    """Scan/verify a QR payload string from the terminal.

    Paste the raw QR payload to verify its signature and check validity.
    """
    if secret is None:
        secret = get_qr_secret()

    valid, data = verify_qr_payload(qr_payload, secret=secret)

    if not valid:
        click.echo(f"✗ INVALID: {data.get('error', 'unknown error')}")
        sys.exit(1)

    pid = data.get("pid", "?")
    name = data.get("nm", "?")
    cat = data.get("cat", "?")
    vf = data.get("vf", "")
    vt = data.get("vt", "")

    now = date.today()
    try:
        d_from = date.fromisoformat(vf)
        d_to = date.fromisoformat(vt)
        in_window = d_from <= now <= d_to
    except (ValueError, TypeError):
        in_window = False

    click.echo(f"  Pass ID:    {pid}")
    click.echo(f"  Name:       {name}")
    click.echo(f"  Category:   {cat}")
    click.echo(f"  Valid:      {vf} → {vt}")
    click.echo(f"  Signature:  ✓ VALID")
    click.echo(f"  Date check: {'✓ ACTIVE' if in_window else '✗ OUTSIDE VALIDITY WINDOW'}")

    if not in_window:
        sys.exit(2)


# ── qr ───────────────────────────────────────────────────────────────────────

@cli.command()
@click.option("-i", "--input", "input_path", required=True,
              type=click.Path(exists=True, path_type=Path),
              help="Pass JSON file to generate QR from.")
@click.option("-o", "--out", default=None, type=click.Path(path_type=Path),
              help="Output PNG path (default: <pass_id>.png in current dir).")
@click.option("--size", default=10, help="QR box size in pixels (default: 10).")
@click.option("--secret", default=None)
@click.option("--transparent/--white-bg", default=True,
              help="Use transparent QR background (default: transparent).")
def qr(input_path: Path, out: Optional[Path], size: int, secret: Optional[str], transparent: bool):
    """Generate a standalone QR code PNG from a pass JSON file."""
    if secret is None:
        secret = get_qr_secret()

    data = json.loads(input_path.read_text())
    gp = GatePass(**data)
    payload = gp.compute_qr_payload(secret=secret)

    out_path = out or Path(f"{gp.pass_id.replace('/', '-').replace(' ', '_')}.png")
    save_qr_image(payload, str(out_path), box_size=size, transparent_bg=transparent, border=0)
    click.echo(f"✓ QR code saved to {out_path}")


# ── template ─────────────────────────────────────────────────────────────────

@cli.command()
@click.option("-o", "--out", default="template.json", type=click.Path(path_type=Path),
              help="Output template file.")
@click.option("--format", "fmt", type=click.Choice(["json", "csv", "xlsx"]), default="json",
              help="Template format.")
def template(out: Path, fmt: str):
    """Generate a sample data template for bulk import."""
    headers = [
        "pass_id", "event_name", "event_type", "category", "full_name",
        "id_number", "phone", "email", "organizer", "valid_from", "valid_to",
        "gate", "table_number", "group_ref", "issued_by",
    ]
    sample_row = [
        "GPX-DINNER-2026-000001", "Annual Gala Dinner", "DINNER", "GUEST",
        "Jane Doe", "REG-001", "+2348012345678", "jane@example.com",
        "Event Co.", "2026-07-01", "2026-07-01", "Main Entrance",
        "T-12", "INV-2026-001", "GatePassX CLI",
    ]

    if fmt == "json":
        sample = {
            "passes": [dict(zip(headers, sample_row))]
        }
        out.write_text(json.dumps(sample, indent=2))
    elif fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Passes"
        ws.append(headers)
        ws.append(sample_row)
        wb.save(str(out))
    else:
        with out.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerow(sample_row)

    click.echo(f"✓ Template written to {out} ({fmt} format)")
    click.echo(f"  Edit the file, then run: gatepassx generate -i {out} -o ./passes")


# ── export-template ──────────────────────────────────────────────────────────

@cli.command("export-template")
@click.option("-t", "--template", "template_path", required=True,
              type=click.Path(exists=True, path_type=Path),
              help="Template PDF to render.")
@click.option("-o", "--out", default=None, type=click.Path(path_type=Path),
              help="Output image path (default: template.png).")
@click.option("--format", "fmt", type=click.Choice(["png", "jpeg", "jpg"]), default="png",
              help="Image format (default: png).")
@click.option("--dpi", default=300, type=int,
              help="Render DPI (default: 300). Use 600+ for print-quality.")
@click.option("--jpeg-quality", default=95, type=int,
              help="JPEG quality 1-100 (default: 95).")
def export_template_cmd(template_path: Path, out: Optional[Path], fmt: str,
                        dpi: int, jpeg_quality: int):
    """Export a template PDF as a high-resolution PNG or JPEG image."""
    ext = "jpg" if fmt in ("jpeg", "jpg") else "png"
    out_path = out or Path(f"{template_path.stem}.{ext}")

    export_template_as_image(
        str(template_path), str(out_path),
        dpi=dpi, fmt=fmt, jpeg_quality=jpeg_quality,
    )
    click.echo(f"✓ Template exported → {out_path} ({fmt.upper()}, {dpi} DPI)")


# ── lock ─────────────────────────────────────────────────────────────────────

@cli.command()
@click.argument("input_file", type=click.Path(exists=True, path_type=Path))
@click.option("-o", "--output-dir", default="generated_passes",
              help="Output directory to record in the lock manifest.")
def lock(input_file: Path, output_dir: str):
    """Lock a data file to prevent alterations after passes are printed.

    Creates a .lock manifest with a SHA-256 hash of the file. Subsequent
    generate commands will refuse if the file has been modified.
    """
    manifest, intact = check_input_lock(str(input_file))
    if manifest is not None and intact:
        click.echo(f"🔒 Already locked: {input_file}")
        click.echo(f"  Generated: {manifest.generated_at.strftime('%Y-%m-%d %H:%M UTC')}")
        click.echo(f"  Passes: {manifest.pass_count}  →  {manifest.output_dir}")
        return

    passes = _load_passes_from_file(input_file)
    m = lock_input_file(str(input_file), output_dir, len(passes))
    click.echo(f"🔒 Locked {input_file.name} ({len(passes)} passes, hash {m.input_hash[:12]}…)")


# ── unlock ───────────────────────────────────────────────────────────────────

@cli.command()
@click.argument("input_file", type=click.Path(path_type=Path))
def unlock(input_file: Path):
    """Remove the lock from a data file, allowing regeneration.

    Use this when you intentionally need to modify the data and regenerate.
    """
    if unlock_input_file(str(input_file)):
        click.echo(f"🔓 Unlocked {input_file.name}")
    else:
        click.echo(f"  No lock file found for {input_file}")


# ── status ───────────────────────────────────────────────────────────────────

@cli.command()
@click.argument("input_file", type=click.Path(exists=True, path_type=Path))
def status(input_file: Path):
    """Check the lock status of a data file."""
    manifest, intact = check_input_lock(str(input_file))
    if manifest is None:
        click.echo(f"  {input_file.name}: unlocked (no lock file)")
        return

    if intact:
        click.echo(f"🔒 {input_file.name}: LOCKED — data intact")
    else:
        click.echo(f"⚠  {input_file.name}: LOCKED — data ALTERED since generation")

    click.echo(f"  Generated: {manifest.generated_at.strftime('%Y-%m-%d %H:%M UTC')}")
    click.echo(f"  Passes:    {manifest.pass_count}")
    click.echo(f"  Output:    {manifest.output_dir}")
    click.echo(f"  Hash:      {manifest.input_hash[:12]}…")
    if not intact:
        current = PassManifest.hash_file(str(input_file))
        click.echo(f"  Current:   {current[:12]}…  ← MISMATCH")
        click.echo(f"  Run 'gatepassx unlock {input_file}' to allow regeneration.")


# ── info ─────────────────────────────────────────────────────────────────────

@cli.command()
@click.option("-i", "--input", "input_path", required=True,
              type=click.Path(exists=True, path_type=Path),
              help="Data file to summarize.")
def info(input_path: Path):
    """Show statistics and summary of a pass data file."""
    passes = _load_passes_from_file(input_path)

    categories: dict = {}
    event_types: dict = {}
    for p in passes:
        cat = p.category.value if hasattr(p.category, "value") else str(p.category)
        et = p.event_type.value if hasattr(p.event_type, "value") else str(p.event_type)
        categories[cat] = categories.get(cat, 0) + 1
        event_types[et] = event_types.get(et, 0) + 1

    click.echo(f"  File:       {input_path}")
    click.echo(f"  Total:      {len(passes)} pass(es)")

    if passes:
        click.echo(f"\n  Event:      {passes[0].event_name}")
        click.echo(f"  Organizer:  {passes[0].organizer}")

    click.echo("\n  By category:")
    for cat, count in sorted(categories.items()):
        click.echo(f"    {cat:<14} {count}")

    click.echo("\n  By event type:")
    for et, count in sorted(event_types.items()):
        click.echo(f"    {et:<14} {count}")

    now = date.today()
    active = sum(1 for p in passes if p.valid_from <= now <= p.valid_to)
    click.echo(f"\n  Active now:  {active} / {len(passes)}")


if __name__ == "__main__":
    cli()
